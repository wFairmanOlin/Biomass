import os
import numpy as np
import csv
import openpyxl
import sys
import tensorflow as tf
np.random.seed(1234)
tf.random.set_seed(35345)
import tensorflow as tf
import keras
from keras.models import Sequential
from keras.layers.core import Dense, Activation, Dropout, Flatten

def weather_mapping(weather):
        maps = {
            'Clear': 0,
            'Light Rain': 0.25,
            'Rain': 0.5,
            'Storm': 0.75
            }
        return maps.get(weather,0)


def preprocess_farmers2018_with_holes(site_number):
    
    file_path = '../Data/farmers_2018/'
    filenames = ["08-02-2018.xlsx", "08-03-2018.xlsx", 
                 "08-04-2018.xlsx", "08-05-2018.xlsx",
                 "08-07-2018.xlsx", "08-08-2018.xlsx",
                 "08-09-2018.xlsx", "08-10-2018.xlsx",
                 "08-11-2018.xlsx", "08-12-2018.xlsx",
                 "08-13-2018.xlsx", "08-14-2018.xlsx",
                 "08-15-2018.xlsx", "08-16-2018.xlsx"]
    
    site1 = []
    weather = []
    first_cell = 'A' + str(3 + site_number)
    last_cell = 'G' + str(3 + site_number)
    daytime = [16, 21, 23, 1, 3, 5]
    daytime_map = [1, 0.3, 0, 0, 0, 0.4]

    
    for file in filenames:

        wb_obj = openpyxl.load_workbook(file_path + file) 
        sheet = wb_obj.active
        # from 6am to 6 pm is night, else day
        # daytimes are 16 - 21 - 23 - 1 - 3 - 5
    
        cells = sheet[first_cell: last_cell]
    
        for c1, c2, c3, c4, c5, c6, c7 in cells:
            
            cell_vals = [c2.value, c3.value, c4.value, c5.value, c6.value, c7.value]
            for i, cell_val in enumerate(cell_vals):
                if '/' in str(cell_val):
                    cell_val_split = str(cell_val).split('/')
                    cell_val_mean = (float(cell_val_split[0]) + float(cell_val_split[1]))/2
                    cell_vals[i] = cell_val_mean
            
            site1.append([c for c in cell_vals])
        
        cells = sheet['A69': 'E69']
        for we, te, th, ts, ws in cells:
            weather.append([te.value, weather_mapping(we.value), ws.value])
              
    #site_1 = [item for sublist in site1 for item in sublist]
    site1_complete = []
    
    for i_day, day in enumerate(site1):
        hour_index = -1
        for measure in day:
            hour_index += 1
            if measure == None or measure == '-':
                continue
            if measure == '15+':
                measure = '16'
            a = [measure]
            for w in weather[i_day]:
                a.append(w)
            #a.append(daytime_map[hour_index])
            a.append(daytime[hour_index])
            site1_complete.append(a)
    
    # data arranged like DO, weather, temperature, wind, 
    # have to normalize all data (in range -1,1, variance1)
    # training on all data - 3

    print('data shape', np.array(site1_complete).shape)
    return np.array(site1_complete).astype(float) 


def normalize(data):
    # print('ds', data.shape)
    normalized = np.zeros(data.shape)
    for i in range(data.shape[1]):    
        dat = data[:, i]
        normalized[:, i] = (dat - dat.mean()) / dat.std()
    
    return normalized, data[:,0].mean(), data[:,0].std()


def normalize_min_max(data):
    # print('ds', data.shape)
    normalized = np.zeros(data.shape)
    for i in range(data.shape[1]):    
        dat = data[:, i]
        normalized[:, i] = (dat - dat.min()) / (dat.max() - dat.min())
    
    return normalized, data[:,0].min(), data[:,0].max()


def normalize_with_time(data):
    # print('ds', data.shape)
    normalized = np.zeros(data.shape)
    for i in range(data.shape[1]-1):    
        dat = data[:, i]
        normalized[:, i] = (dat - dat.mean()) / dat.std()
    
    return normalized, data[:,0].mean(), data[:,0].std()

def normalize_all_do(data):
    data_no_nan = []
    for site in data:
        for cell in site:
            # print(cell)
            if not np.isnan(cell):
                data_no_nan.append(cell)
                
    dat = np.array(data_no_nan).astype(float)
    return (data - dat.mean()) / dat.std(), dat.mean(), dat.std() 


def remove_nan(data):
    data_clean = data
    deleted = 0
    for index, do_val in enumerate(data[0, :]):
        if np.isnan(do_val):
            data_clean = np.delete(data_clean, index - deleted, axis=1)
            deleted += 1               
    
    return data_clean

def normalize_other(data):
    return (data - data.mean()) / data.std(), data.mean(), data.std() 

def feed_forward_nn():
    
    model = Sequential()
    model.add(Dense(units=1, input_dim=4, kernel_initializer="uniform")) # std input = id, time, weather, (can add Temp, th, ts, timapping, more specific timeday)
    model.add(Activation("tanh"))
    optimizer = keras.optimizers.Adam(learning_rate=0.001) 
    model.compile(loss="mean_squared_error", optimizer=optimizer)
    
    return model


def fill_nan(encoder, X):
    x_filled = X[0, :]
    #print(np.expand_dims(X[1:, 1], axis=0).shape)
    for index, do_val in enumerate(x_filled):
        if np.isnan(do_val):
            do_predicted = encoder.predict(np.expand_dims(X[1:, index], axis=0))
            x_filled[index] = do_predicted
    
    return x_filled


def preprocess_dataset_filling_holes():
    
    file_path = '../Data/farmers_2018/'
    filenames = ["08-02-2018.xlsx", "08-03-2018.xlsx", 
                 "08-04-2018.xlsx", "08-05-2018.xlsx",
                 "08-07-2018.xlsx", "08-08-2018.xlsx",
                 "08-09-2018.xlsx", "08-10-2018.xlsx",
                 "08-11-2018.xlsx", "08-12-2018.xlsx",
                 "08-13-2018.xlsx", "08-14-2018.xlsx",
                 "08-15-2018.xlsx", "08-16-2018.xlsx"]

    columns_per_sheet = 6
    sheet_number = len(filenames)
    total_length = columns_per_sheet*sheet_number
    sites_number = 65   # + 1 for the daytime
    do_data = np.zeros((sites_number, total_length))
    other_data = []
    daytime = [16, 21, 23, 1, 3, 5]
    daytime_map = [1, 0.3, 0, 0, 0, 0.4]
    
    for file_index, file in enumerate(filenames):
        
        wb_obj = openpyxl.load_workbook(file_path + file) 
        ws = wb_obj.active
        day_other_data = []
    
        for i_row, row in enumerate(ws.iter_rows(4, 69)):
            
            i_length = file_index * columns_per_sheet
            
            for i_cell, cell in enumerate(row):
                if i_row != 65:  # last row is for weather
                    if i_cell != 0: # no need for pond index
                        
                        if cell.value == '15+':
                            cell.value = 16
                        elif cell.value == 'None' or cell.value == '-':
                            cell.value = float('NaN')
                        elif '/' in str(cell.value):
                            cell_val_split = str(cell.value).split('/')
                            cell.value = (float(cell_val_split[0]) + float(cell_val_split[1]))/2
                            
                        do_data[i_row, i_length + i_cell-1] = cell.value
                        #print(cell.value, i_row)
                else:
                    if i_cell == 0:
                        day_other_data.append(weather_mapping(cell.value))
                    elif i_cell == 1 or i_cell == 4:  # we te wi
                        day_other_data.append(cell.value)
                    
        other_data.append(day_other_data)
                
        
    do_data_norm, do_mean, do_std = normalize_all_do(do_data)
    print('data shape', do_data_norm.shape)
    
    time = np.expand_dims(np.array(daytime), axis=0)
    times = (np.repeat(a=time, repeats=sheet_number ,axis=0)).flatten()
    time_n, t_mean, t_std = normalize_other(times)
    time_norm = np.expand_dims(time_n, axis=0)
    
    other_datas = np.array(other_data).astype(float)
    weathers = np.repeat(a=other_datas[:,0], repeats=columns_per_sheet, axis=0)
    weather_n, we_mean, we_std = normalize_other(weathers)
    weather_norm = np.expand_dims(weather_n, axis=0)

    temperatures = np.repeat(a=other_datas[:,1], repeats=columns_per_sheet, axis=0)
    temp_n, temp_mean, temp_std = normalize_other(temperatures)
    temp_norm = np.expand_dims(temp_n, axis=0)
    
    winds = np.repeat(a=other_datas[:,2], repeats=columns_per_sheet, axis=0)
    winds_n, wi_mean, wi_std = normalize_other(winds)
    winds_norm = np.expand_dims(winds_n, axis=0)
    
    encoder = feed_forward_nn()
    
    rmse = []
    mape = []
    
    for n_site in range(sites_number):
        do_data_n = np.expand_dims(do_data_norm[n_site,:], axis=0)
        X_with_nan = np.concatenate((do_data_n, time_norm, weather_norm, temp_norm, winds_norm), axis=0)
        X = remove_nan(X_with_nan)
     
        x_train = np.transpose(X[1:, :-4])
        y_train = X[0, :-4]
        
        x_test = np.transpose(X[1:, -4:])
        y_test = X[0, -4:]
            
        print('xtrain', x_train.shape, 'ytrain', y_train.shape)
        encoder.fit(x_train, y_train, batch_size=1, epochs=500)
        print('xtest', x_test.shape)
        y_pred = encoder.predict(x_test)
        
        y_pred_d = y_pred * do_std + do_mean
        y_test_d = y_test * do_std + do_mean
        # should denormalize
        
        rmse.append(np.sqrt((np.mean((np.absolute(y_pred_d - y_test_d)) ** 2))))
        mape.append(1 - np.mean(np.abs(np.true_divide(y_test_d - y_pred_d, y_test_d))))
        
        # save model        
        encoder.save('models/encoder_model.h5')
        
        # predict the Nan and return the dataset
        do_data_site_filled = fill_nan(encoder, X_with_nan)
        do_data[n_site, :] = do_data_site_filled
        
    do_data = do_data*do_std + do_mean
    
    print('mape', np.array(mape).mean(), 'rmse', np.array(rmse).mean())
    
    return do_data


#data = preprocess_dataset_filling_holes()
#np.savetxt('data_filled/data_farmers_2018.csv', data, delimiter=' ')

def get_data_filled(site):
    data_tot = np.genfromtxt(os.path.join('data_filled', 'data_farmers_2018.csv'))
    return data_tot[site, :]


def c_map(condition):
        maps = {
            'Fair': 0,
            'Fog': 0.10,
            'Partly Cloudy': 0.25,
            'Mostly Cloudy': 0.5,
            'Cloudy': 0.75
            }
        return maps.get(condition,0)
    

def wd_map(wd):
        maps = {
            'CALM': 0,
            'SE':315,
            'S':270,
            'SSW':247.5,
            'SW':225,
            'W':180,
            'NW':135,
            'N':90
            }
        return maps.get(wd,0)


def preprocess_updated_dataset_with_holes(site_number):
    
    file_path = '../Data/farmers_2018_weather_updated/'
    filenames = ["08-02-2018.xlsx", "08-03-2018.xlsx", 
                 "08-04-2018.xlsx", "08-05-2018.xlsx",
                 "08-07-2018.xlsx", "08-08-2018.xlsx",
                 "08-09-2018.xlsx", "08-10-2018.xlsx",
                 "08-11-2018.xlsx", "08-12-2018.xlsx",
                 "08-13-2018.xlsx", "08-14-2018.xlsx",
                 "08-15-2018.xlsx", "08-16-2018.xlsx"]
    
    site1 = []
    temp = []
    cond = []
    ws = []
    wd = []
    first_cell = 'A' + str(3 + site_number)
    last_cell = 'G' + str(3 + site_number)
    
    for file in filenames:

        nan_per_site = 0
        
        wb_obj = openpyxl.load_workbook(file_path + file) 
        sheet = wb_obj.active
        # from 6am to 6 pm is night, else day
        # daytimes are 16 - 21 - 23 - 1 - 3 - 5
    
        cells = sheet[first_cell: last_cell]
    
        for c1, c2, c3, c4, c5, c6, c7 in cells:
            
            cell_vals = [c2.value, c3.value, c4.value, c5.value, c6.value, c7.value]
            for i, cell_val in enumerate(cell_vals):
                if '/' in str(cell_val):
                    cell_val_split = str(cell_val).split('/')
                    cell_val_mean = (float(cell_val_split[0]) + float(cell_val_split[1]))/2
                    cell_vals[i] = cell_val_mean
            
            site1.append([c for c in cell_vals])
        
        t_cells = sheet['B70': 'G70']
        for t1, t2, t3, t4, t5, t6 in t_cells: 
            temp.append([t1.value, t2.value, t3.value, t4.value, t5.value, t6.value])
            
        cond_cells = sheet['B72': 'G72']
        for c1, c2, c3, c4, c5, c6 in cond_cells:
            cond.append([c_map(c1.value), c_map(c2.value), c_map(c3.value),
                         c_map(c4.value), c_map(c5.value), c_map(c6.value)])
        
        #ws_cells = sheet['B73': 'G73']
        #for ws1, ws2, ws3, ws4, ws5, ws6 in ws_cells: 
        #    ws.append([ws1.value, ws2.value, ws3.value, ws4.value, ws5.value, ws6.value])
        
        #wd_cells = sheet['B74': 'G74']
        #for wd1, wd2, wd3, wd4, wd5, wd6 in wd_cells:
        #    wd.append([wd_map(wd1.value), wd_map(wd2.value), wd_map(wd3.value),
        #                 wd_map(wd4.value), wd_map(wd5.value), wd_map(wd6.value)])
        
            
    #site_1 = [item for sublist in site1 for item in sublist]
    site1_complete = []
    for i_day, day in enumerate(site1):
        for i_m, measure in enumerate(day):
            if measure == None or measure == '-':
                nan_per_site += 1
                continue
            if measure == '15+':
                measure = '16'
            a = [measure, temp[i_day][i_m], cond[i_day][i_m]]
            # , ws[i_day][i_m], wd[i_day][i_m]
            site1_complete.append(a)
    
    # data arranged like DO, weather, temperature, wind, 
    # have to normalize all data (in range -1,1, variance1)
    # training on all data - 3

    # print('nan', site_number, ':', nan_per_site)
    return np.array(site1_complete).astype(float) 
 


